import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Load data
df = pd.read_csv("exporters/reports/all_sites_last30days.csv")

# Create an Excel writer
excel_path = "exporters/reports/advanced_dashboard.xlsx"
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Export data to Excel
    df.to_excel(writer, sheet_name='Metrics Data', index=False)

    # Get workbook and sheet
    workbook = writer.book
    sheet = workbook['Metrics Data']

    # Create Pivot Table for Metrics by Site
    pivot_table_sheet = workbook.create_sheet('Pivot Table')
    pivot_table_data = df.pivot_table(index='site', values=['site_packet_loss_percent', 'site_latency_ms', 'site_cpu_usage_percent'],
                                      aggfunc='mean')
    for row in dataframe_to_rows(pivot_table_data, index=True, header=True):
        pivot_table_sheet.append(row)
    
    # Adding a bar chart for Packet Loss
    chart_packet_loss = BarChart()
    chart_packet_loss.title = "Packet Loss by Site"
    data_packet_loss = Reference(pivot_table_sheet, min_col=2, min_row=1, max_col=2, max_row=len(pivot_table_data)+1)
    categories_packet_loss = Reference(pivot_table_sheet, min_col=1, min_row=2, max_row=len(pivot_table_data)+1)
    chart_packet_loss.add_data(data_packet_loss, titles_from_data=True)
    chart_packet_loss.set_categories(categories_packet_loss)
    pivot_table_sheet.add_chart(chart_packet_loss, "E5")

    # Add Conditional Formatting for High Latency
    for row in range(2, len(df) + 2):
        cell = sheet.cell(row=row, column=3)  # Latency column
        if cell.value > 100:  # Highlight if latency > 100ms
            cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # Create line chart for Latency over time
    chart_latency = LineChart()
    chart_latency.title = "Latency (ms) by Site"
    data_latency = Reference(sheet, min_col=3, min_row=1, max_col=3, max_row=len(df)+1)
    categories_latency = Reference(sheet, min_col=1, min_row=2, max_row=len(df)+1)
    chart_latency.add_data(data_latency, titles_from_data=True)
    chart_latency.set_categories(categories_latency)
    sheet.add_chart(chart_latency, "E20")

    # Create pie chart for CPU Usage by Site
    chart_cpu = PieChart()
    chart_cpu.title = "CPU Usage by Site"
    data_cpu = Reference(sheet, min_col=4, min_row=1, max_col=4, max_row=len(df)+1)
    categories_cpu = Reference(sheet, min_col=1, min_row=2, max_row=len(df)+1)
    chart_cpu.add_data(data_cpu, titles_from_data=True)
    chart_cpu.set_categories(categories_cpu)
    sheet.add_chart(chart_cpu, "E35")

# Save Excel file
print(f"Excel dashboard saved at {excel_path}")
